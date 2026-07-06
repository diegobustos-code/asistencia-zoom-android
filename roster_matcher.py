# -*- coding: utf-8 -*-
"""
roster_matcher.py
==================
Módulo encargado de cotejar el listado OFICIAL de socios (un Excel con las
columnas Sede, Apellido Paterno, Apellido Materno, Nombres) contra los
participantes leídos desde el CSV de Zoom, para producir un listado de
asistencia basado en la lista oficial: una fila POR CADA SOCIO, marcado
como "P" (Presente) o "A" (Ausente) según si se le encontró y cuántos
minutos acumuló en la reunión de Zoom.

Por qué existe este módulo:
    El nombre que la gente escribe en Zoom es poco confiable (apodos,
    sedes pegadas al nombre, orden invertido, apellidos incompletos, etc.).
    En cambio, el listado oficial de socios es la fuente de verdad. Este
    módulo usa el listado oficial como base y busca, para cada socio, si
    hay algún participante de Zoom cuyo nombre coincida razonablemente.

Cómo funciona el cruce (en resumen):
    1. Se "tokeniza" (separa en palabras normalizadas, sin tildes ni
       mayúsculas) tanto el nombre de cada socio oficial como el nombre de
       cada participante de Zoom.
    2. Para que un socio sea candidato a coincidir con un participante de
       Zoom, al menos su apellido paterno o su apellido materno debe
       aparecer entre las palabras del nombre de Zoom.
    3. Entre los candidatos, se elige el que tiene más palabras en común
       (con más peso si el apellido paterno o materno coincide completo).
       Si hay empate exacto entre dos o más candidatos, ese participante
       de Zoom se marca como "ambiguo" y NO se asigna automáticamente
       (para evitar adjudicar minutos de asistencia a la persona
       equivocada); queda disponible para revisión manual.
    4. Si un participante de Zoom no coincide con ningún socio, se marca
       como "no encontrado" (puede ser un invitado, un error de tipeo muy
       grande, o alguien que no está en el listado oficial).
    5. Si un mismo socio aparece más de una vez en el Zoom (por ejemplo,
       se desconectó y volvió a entrar), sus minutos se SUMAN.

Nada de esto es 100% infalible con datos tan heterogéneos como los
nombres escritos a mano en Zoom, por lo que el resultado incluye listas
separadas de "no encontrados" y "ambiguos" para que se puedan revisar y
corregir manualmente si hace falta.
"""

import re
import unicodedata
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from csv_processor import compute_attendance_flag


class RosterProcessingError(Exception):
    """Error controlado y con mensaje claro para mostrar al usuario final."""
    pass


ROSTER_COLUMNS = ["Sede", "Apellido Paterno", "Apellido Materno", "Nombres"]

# Palabras que se ignoran al comparar nombres (no aportan para identificar
# a una persona: cargos, dispositivos, artículos, restos de "sede", etc.)
STOPWORDS = {
    "sede", "cede", "prof", "profesor", "profesora", "srta", "sra", "sr",
    "don", "dona", "doña", "usuario", "zoom", "invitado", "guest",
    "de", "del", "la", "el", "los", "las", "y",
    "iphone", "ipad", "galaxy", "samsung", "xiaomi", "huawei", "android",
    "moto", "macbook", "note", "ultra", "pro", "plus", "mini", "windows",
    "pc", "tablet", "of",
}


def _normalize(text: str) -> str:
    text = (text or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def tokenize(text: str) -> List[str]:
    """Separa un texto en palabras normalizadas útiles para comparar nombres."""
    normalized = _normalize(text)
    normalized = re.sub(r"[^a-z\s]", " ", normalized)
    return [t for t in normalized.split() if len(t) >= 2 and t not in STOPWORDS]


def load_roster_excel(filepath: str) -> List[Dict[str, str]]:
    """
    Carga el listado oficial de socios desde un archivo Excel (.xlsx) con
    las columnas: Sede, Apellido Paterno, Apellido Materno, Nombres
    (el orden de las columnas no importa, se detectan por nombre).

    Lanza RosterProcessingError con un mensaje claro si falta alguna
    columna obligatoria o el archivo está vacío.
    """
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        raise RosterProcessingError(
            f"No se pudo abrir el archivo Excel del listado de socios.\n"
            f"Detalle técnico: {e}"
        )

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RosterProcessingError("El listado de socios está vacío.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_map = {}
    for wanted in ROSTER_COLUMNS:
        wanted_norm = _normalize(wanted)
        match = None
        for h in headers:
            if _normalize(h) == wanted_norm:
                match = h
                break
        header_map[wanted] = match

    missing = [col for col, real in header_map.items() if real is None]
    if missing:
        raise RosterProcessingError(
            "El listado de socios no tiene las siguientes columnas "
            "obligatorias:\n" + "\n".join(f"  • {m}" for m in missing)
            + "\n\nEncabezados encontrados en el archivo:\n"
            + ", ".join(h for h in headers if h)
            + "\n\nLas columnas deben llamarse exactamente: "
            + ", ".join(ROSTER_COLUMNS)
        )

    col_index = {col: headers.index(real) for col, real in header_map.items()}

    roster: List[Dict[str, str]] = []
    for row in rows[1:]:
        if row is None or not any(row):
            continue
        record = {
            col: str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""
            for col, idx in col_index.items()
        }
        # Ignorar filas completamente vacías de nombre
        if not record["Apellido Paterno"] and not record["Apellido Materno"] and not record["Nombres"]:
            continue
        roster.append(record)

    if not roster:
        raise RosterProcessingError(
            "El listado de socios se leyó correctamente, pero no se "
            "encontraron filas de socios válidas."
        )

    return roster


def _score(entry: Dict[str, object], zoom_tokens: set) -> Optional[Tuple[int, float]]:
    """
    Calcula qué tan bien coincide un socio (`entry`) con las palabras del
    nombre de un participante de Zoom (`zoom_tokens`). Devuelve None si ni
    siquiera el apellido paterno ni el materno aparecen (candidato
    descartado de plano). Si hay coincidencia, devuelve una tupla
    (puntaje_bruto, similitud_jaccard) usada para ordenar candidatos.
    """
    paterno: set = entry["_paterno"]
    materno: set = entry["_materno"]
    all_tokens: set = entry["_all"]

    if not (paterno & zoom_tokens) and not (materno & zoom_tokens):
        return None

    overlap = all_tokens & zoom_tokens
    raw_score = len(overlap)
    if paterno and paterno.issubset(zoom_tokens):
        raw_score += 2
    if materno and materno.issubset(zoom_tokens):
        raw_score += 1

    union = all_tokens | zoom_tokens
    jaccard = len(overlap) / len(union) if union else 0.0
    return (raw_score, jaccard)


def match_zoom_to_roster(
    zoom_records: List[Dict[str, object]],
    roster_records: List[Dict[str, str]],
    attendance_threshold: float = 30.0,
) -> Tuple[List[Dict[str, object]], List[Dict[str, object]], List[Dict[str, object]]]:
    """
    Coteja los participantes de Zoom contra el listado oficial de socios.

    Devuelve una tupla (resultado, no_encontrados, ambiguos):

      - resultado: una fila POR CADA SOCIO del listado oficial, con las
        llaves "Nombre", "Apellido", "Sede", "Duración (minutos)" y
        "Asistencia" (P/A), lista para mostrarse en la tabla / exportarse
        igual que el resto de la aplicación.

      - no_encontrados: participantes de Zoom que no coincidieron con
        ningún socio del listado oficial (para revisión manual).

      - ambiguos: participantes de Zoom que coincidieron por igual con
        dos o más socios, por lo que no se pudo determinar automáticamente
        a cuál asignarle la asistencia (para revisión manual).
    """
    # Preparar los socios con sus conjuntos de palabras (tokens)
    entries = []
    for r in roster_records:
        paterno = set(tokenize(r["Apellido Paterno"]))
        materno = set(tokenize(r["Apellido Materno"]))
        nombres = set(tokenize(r["Nombres"]))
        entries.append({
            **r,
            "_paterno": paterno,
            "_materno": materno,
            "_nombres": nombres,
            "_all": paterno | materno | nombres,
            "_duracion_total": 0.0,
            "_matched_zoom_names": [],
        })

    no_encontrados: List[Dict[str, object]] = []
    ambiguos: List[Dict[str, object]] = []

    for zr in zoom_records:
        zoom_tokens = set(tokenize(str(zr.get("Nombre", ""))))
        candidates = []
        for idx, entry in enumerate(entries):
            s = _score(entry, zoom_tokens)
            if s is not None:
                candidates.append((s, idx))

        if not candidates:
            no_encontrados.append(zr)
            continue

        candidates.sort(key=lambda c: c[0], reverse=True)
        top_score = candidates[0][0]
        top_matches = [idx for s, idx in candidates if s == top_score]

        if len(top_matches) > 1:
            ambiguos.append(zr)
            continue

        idx = top_matches[0]
        entries[idx]["_duracion_total"] += float(zr.get("Duración (minutos)", 0) or 0)
        entries[idx]["_matched_zoom_names"].append(str(zr.get("Nombre", "")))

    resultado: List[Dict[str, object]] = []
    for entry in entries:
        duracion = round(entry["_duracion_total"], 2)
        apellido = f"{entry['Apellido Paterno']} {entry['Apellido Materno']}".strip()
        apellido = re.sub(r"\s{2,}", " ", apellido)
        resultado.append({
            "Nombre": entry["Nombres"].strip(),
            "Apellido": apellido,
            "Sede": entry["Sede"].strip(),
            "Duración (minutos)": duracion,
            "Asistencia": compute_attendance_flag(duracion, attendance_threshold),
        })

    return resultado, no_encontrados, ambiguos
