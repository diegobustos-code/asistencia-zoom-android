# -*- coding: utf-8 -*-
"""
excel_exporter.py
==================
Exporta la lista de participantes filtrados a un archivo .xlsx con formato
de Tabla de Excel (con filtros automáticos habilitados), estilo aplicado,
columnas numéricas para la duración, columna de Asistencia (P/A) con color
tipo semáforo, y ancho de columna ajustado automáticamente.
"""

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUMNS = ["Nombre", "Apellido", "Sede", "Duración (minutos)", "Asistencia"]

# Colores estilo "semáforo" para la columna Asistencia (P = Presente, A = Ausente)
_FILL_PRESENTE = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
_FONT_PRESENTE = Font(color="FF006100", bold=True)
_FILL_AUSENTE = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
_FONT_AUSENTE = Font(color="FF9C0006", bold=True)


def export_to_excel(records: List[Dict[str, object]], filepath: str,
                     sheet_name: str = "Asistencia") -> None:
    """
    Crea un archivo .xlsx a partir de `records` con las columnas
    Nombre, Apellido, Sede, Duración (minutos) y Asistencia (P/A),
    formateado como una Tabla de Excel real (no solo un rango), con
    filtros automáticos, estilo TableStyleMedium9, colores tipo semáforo
    para la columna Asistencia y anchos de columna ajustados.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limita el nombre de hoja a 31 caracteres

    # --- Encabezados ---
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # --- Filas de datos ---
    duration_col_index = COLUMNS.index("Duración (minutos)") + 1
    attendance_col_index = COLUMNS.index("Asistencia") + 1

    for record in records:
        duration_value = record.get("Duración (minutos)", 0)
        try:
            duration_value = float(duration_value)
        except (TypeError, ValueError):
            duration_value = 0.0

        row = [
            record.get("Nombre", ""),
            record.get("Apellido", ""),
            record.get("Sede", ""),
            duration_value,
            record.get("Asistencia", ""),
        ]
        ws.append(row)

    last_row = ws.max_row
    last_col_letter = get_column_letter(len(COLUMNS))

    # --- Convertir el rango en una Tabla de Excel real con filtros ---
    if last_row >= 1:
        table_range = f"A1:{last_col_letter}{last_row}"
        table = Table(displayName="TablaAsistencia", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # --- Formato numérico para la columna de duración ---
    for row_idx in range(2, last_row + 1):
        cell = ws.cell(row=row_idx, column=duration_col_index)
        cell.number_format = "0.00"

    # --- Color tipo semáforo + centrado para la columna Asistencia ---
    for row_idx in range(2, last_row + 1):
        cell = ws.cell(row=row_idx, column=attendance_col_index)
        cell.alignment = Alignment(horizontal="center")
        if cell.value == "P":
            cell.fill = _FILL_PRESENTE
            cell.font = _FONT_PRESENTE
        elif cell.value == "A":
            cell.fill = _FILL_AUSENTE
            cell.font = _FONT_AUSENTE

    # --- Ajuste automático de ancho de columnas ---
    for col_index, column_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_index)
        max_length = len(str(column_name))
        for row_idx in range(2, last_row + 1):
            value = ws.cell(row=row_idx, column=col_index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col_letter].width = max_length + 4

    ws.freeze_panes = "A2"
    wb.save(filepath)
