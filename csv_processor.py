# -*- coding: utf-8 -*-
"""
csv_processor.py
=================
Módulo responsable de leer y normalizar los archivos CSV de participantes
exportados desde Zoom.

Funciones principales:
    - load_zoom_csv(filepath): lee el CSV, detecta las columnas relevantes
      (sin importar el orden en que vengan) y devuelve una lista de
      diccionarios homogénea con las llaves:
          "Nombre", "Apellido", "Sede", "Duración (minutos)"

    - export_records_to_csv(records, filepath): exporta la lista filtrada
      a un nuevo archivo CSV.

Si en el futuro Zoom cambia el nombre de las columnas en su exportación,
NO es necesario tocar la lógica del programa: basta con agregar las nuevas
palabras clave en las listas KEYWORDS_* que están al inicio de este archivo.
Ver la sección "Cómo adaptar el programa a cambios de Zoom" en el README.
"""

import csv
import io
import re
import unicodedata
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


class CSVProcessingError(Exception):
    """Error controlado y con mensaje claro para mostrar al usuario final."""
    pass


# ---------------------------------------------------------------------------
# PALABRAS CLAVE PARA DETECCIÓN AUTOMÁTICA DE COLUMNAS
# ---------------------------------------------------------------------------
# El orden dentro de cada lista importa: las primeras tienen mayor prioridad
# cuando hay coincidencias exactas con el encabezado normalizado (minúsculas
# y sin tildes). Agrega aquí nuevas variantes si Zoom cambia sus encabezados.
KEYWORDS_FULL_NAME = [
    "name (original name)",
    "nombre (nombre original)",
    "nombre completo",
    "participant name",
    "nombre del participante",
    "attendee name",
    "attendee",
    "participante",
    "name",
    "nombre",
]
KEYWORDS_FIRST_NAME = ["first name", "primer nombre", "nombre de pila"]
KEYWORDS_LAST_NAME = ["last name", "apellidos", "apellido"]
KEYWORDS_SEDE = [
    "sede", "ubicacion", "ubicación", "location", "site",
    "sala", "room", "oficina", "office", "sucursal", "campus",
]
KEYWORDS_DURATION = [
    "duration (minutes)",
    "duracion (minutos)",
    "duración (minutos)",
    "time in session (minutes)",
    "tiempo en sesion (minutos)",
    "total duration (minutes)",
    "attendance duration",
    "duration",
    "duracion",
    "duración",
    "minutes",
    "minutos",
    "tiempo",
]

# Encabezados que NO deben tomarse como "nombre" aunque contengan "name",
# para evitar falsos positivos como "User Email" o "Screen Name ID".
EXCLUDE_NAME_HINTS = ["email", "correo", "id", "guest", "invitado"]

# Columnas requeridas como mínimo para poder trabajar con el archivo.
REQUIRED_FIELDS_HINT = "Nombre del participante y Duración de asistencia"


# ---------------------------------------------------------------------------
# LISTA DE SEDES CONOCIDAS (para organizaciones que embeben la sede dentro
# del campo de nombre en el CSV de Zoom, en formatos muy variados).
# ---------------------------------------------------------------------------
# Si tu organización no usa estas sedes, o usa otras, edita este diccionario:
# clave = nombre "bonito" que se mostrará en la columna Sede,
# valor = lista de variantes (en minúsculas y SIN tildes) que pueden aparecer
#         en el nombre del participante para identificar esa sede.
# El emparejamiento es insensible a mayúsculas/tildes y usa límites de
# palabra, por lo que no hace falta listar variantes con mayúsculas.
KNOWN_SEDES: Dict[str, List[str]] = {
    "Arica": ["arica"],
    "Alto Hospicio": ["alto hospicio"],
    "Iquique": ["iquique"],
    "Antofagasta": ["antofagasta"],
    "Copiapó": ["copiapo"],
    "La Serena": ["la serena"],
    "Coquimbo": ["coquimbo"],
    "Valparaíso": ["valparaiso"],
    "Viña del Mar": ["vina del mar"],
    "Quilpué": ["quilpue"],
    "Quillota": ["quillota"],
    "Rancagua": ["rancagua"],
    "Talca": ["talca"],
    "Curicó": ["curico"],
    "San Fernando": ["san fernando"],
    "Chillán": ["chillan"],
    "Concepción": ["concepcion"],
    "Temuco": ["temuco"],
    "Valdivia": ["valdivia"],
    "Osorno": ["osorno"],
    "Puerto Montt": ["puerto montt", "pto montt", "pto. montt", "ptomontt"],
    "Puente Alto": ["puente alto", "pte. alto", "pte alto"],
    "Alameda": ["alameda"],
    "Independencia": ["independencia", "independecia"],
    "Maipú": ["maipu"],
    "San Bernardo": ["san bernardo", "san bdo", "sn bdo", "sn. bdo.", "sn. bdo"],
    "La Florida": ["la florida", "la floridda"],
    "Ñuñoa": ["nunoa", "nuloa"],
    "Las Condes": ["las condes"],
    "Providencia": ["providencia"],
    "Gran Avenida": ["gran avenida"],
    "Quilicura": ["quilicura"],
    "Los Ángeles": ["los angeles"],
}

# Lista aplanada (variante, nombre_bonito), ordenada de más larga a más corta
# para que las coincidencias de varias palabras (ej. "puerto montt") tengan
# prioridad sobre coincidencias parciales de una sola palabra.
_SEDE_VARIANTS: List[Tuple[str, str]] = sorted(
    ((variant, canonical) for canonical, variants in KNOWN_SEDES.items() for variant in variants),
    key=lambda vc: len(vc[0]),
    reverse=True,
)

_SEDE_KEYWORD_PATTERN = re.compile(r"(sede|cede)\b", re.IGNORECASE)


def _normalize(text: str) -> str:
    """Minúsculas y sin tildes, para comparar encabezados de forma robusta."""
    text = text.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def _find_column(headers: List[str], keywords: List[str],
                  exclude_hints: Optional[List[str]] = None) -> Optional[str]:
    """
    Busca en `headers` la columna que mejor coincide con `keywords`.
    Primero intenta coincidencia exacta (encabezado normalizado == keyword),
    y si no encuentra nada, intenta coincidencia parcial (keyword contenida
    dentro del encabezado).
    """
    exclude_hints = exclude_hints or []
    normalized = {h: _normalize(h) for h in headers}

    for kw in keywords:
        for original, norm in normalized.items():
            if norm == kw and not any(ex in norm for ex in exclude_hints):
                return original

    for kw in keywords:
        for original, norm in normalized.items():
            if kw in norm and not any(ex in norm for ex in exclude_hints):
                return original

    return None


def detect_columns(headers: List[str]) -> Dict[str, Optional[str]]:
    """Devuelve un diccionario con el nombre real de cada columna detectada."""
    return {
        "full_name": _find_column(headers, KEYWORDS_FULL_NAME, EXCLUDE_NAME_HINTS),
        "first_name": _find_column(headers, KEYWORDS_FIRST_NAME),
        "last_name": _find_column(headers, KEYWORDS_LAST_NAME),
        "sede": _find_column(headers, KEYWORDS_SEDE),
        "duration": _find_column(headers, KEYWORDS_DURATION),
    }


def _read_text_with_fallback_encoding(filepath: str) -> str:
    """Lee el archivo probando UTF-8 (con/sin BOM), luego Windows-1252 y Latin-1."""
    encodings_to_try = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    last_error: Optional[Exception] = None
    for enc in encodings_to_try:
        try:
            with open(filepath, "r", encoding=enc, newline="") as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError) as e:
            last_error = e
            continue
    raise CSVProcessingError(
        "No fue posible leer el archivo con las codificaciones soportadas "
        f"(UTF-8 / Windows-1252 / Latin-1).\nDetalle técnico: {last_error}"
    )


def _load_disguised_excel_as_csv_text(filepath: str) -> str:
    """
    Algunos flujos de descarga de reportes de Zoom (o el propio Excel al
    "Guardar como CSV" en ciertas configuraciones) generan un archivo con
    extensión .csv que en realidad es un archivo Excel (.xlsx) por dentro.
    Esta función detecta y lee ese caso, reconstruyendo el texto CSV
    equivalente para que pueda procesarse con la misma lógica que un CSV
    normal.
    """
    with open(filepath, "rb") as f:
        data = f.read()

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise CSVProcessingError(
            "El archivo parece ser un Excel (.xlsx) por dentro, pero no se "
            f"pudo abrir correctamente.\nDetalle técnico: {e}"
        )

    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        raise CSVProcessingError("El archivo está vacío.")

    max_col = max(len(r) for r in raw_rows)
    buffer = io.StringIO()
    if max_col <= 1:
        # Cada fila del Excel contiene, en una sola celda, una línea
        # completa de texto CSV (encabezado y datos separados por comas).
        for r in raw_rows:
            value = r[0] if r else None
            buffer.write(("" if value is None else str(value)) + "\n")
    else:
        # El Excel ya viene con los datos repartidos en columnas reales.
        writer = csv.writer(buffer)
        for r in raw_rows:
            writer.writerow(["" if v is None else v for v in r])

    return buffer.getvalue()


def is_disguised_excel_file(filepath: str) -> bool:
    """Versión pública de `_is_disguised_excel_file`, para que la interfaz
    gráfica pueda informarle al usuario si su ".csv" es en realidad un
    Excel por dentro."""
    return _is_disguised_excel_file(filepath)


def _is_disguised_excel_file(filepath: str) -> bool:
    """Detecta si el archivo es en realidad un .xlsx (ZIP/Office Open XML)
    sin importar su extensión, revisando su firma binaria."""
    try:
        with open(filepath, "rb") as f:
            signature = f.read(4)
    except OSError:
        return False
    return signature == b"PK\x03\x04"


def _sniff_delimiter(sample: str) -> str:
    """Detecta el delimitador (coma, punto y coma o tabulación). Por defecto: coma."""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        return ","


def _normalize_keep_length(text: str) -> str:
    """
    Igual que `_normalize`, pero garantiza que el resultado tenga la MISMA
    longitud que el texto original (carácter a carácter), para poder mapear
    las posiciones de una coincidencia de vuelta al texto original sin
    desalinearse por culpa de tildes/eñes.
    """
    chars = []
    for ch in text:
        decomposed = unicodedata.normalize("NFKD", ch)
        base = "".join(c for c in decomposed if not unicodedata.combining(c))
        chars.append((base if base else ch)[:1].lower() or ch.lower())
    return "".join(chars)


def _find_known_sede(normalized_text: str) -> Optional[Tuple[str, int, int]]:
    """Busca la primera sede conocida (de KNOWN_SEDES) dentro del texto ya
    normalizado. Devuelve (nombre_bonito, inicio, fin) o None."""
    for variant, canonical in _SEDE_VARIANTS:
        pattern = r"(?<![a-z0-9])" + re.escape(variant) + r"(?![a-z0-9])"
        match = re.search(pattern, normalized_text)
        if match:
            return canonical, match.start(), match.end()
    return None


def extract_name_and_sede(raw_name: str) -> Tuple[str, str]:
    """
    Separa el nombre del participante de la sede, cuando ambos vienen
    mezclados en el mismo campo (algo común en exportaciones de Zoom donde
    la organización no configuró un campo de "sede" y la gente la escribe
    manualmente junto a su nombre, en formatos muy variados).

    Estrategia (en este orden):
      1. Comparar el texto completo contra la lista de sedes conocidas
         (KNOWN_SEDES), sin importar si viene acompañada o no de la palabra
         "sede". Esto es lo más confiable porque se basa en nombres reales
         de sedes/ciudades y no en adivinar patrones de texto.
      2. Si no hay coincidencia con la lista, pero el texto contiene la
         palabra "sede"/"cede" seguida de texto (letras), se usa ese texto
         como sede (para no perder información en organizaciones con sedes
         no listadas en KNOWN_SEDES).
      3. Si no se detecta nada, se deja la sede en blanco y se conserva el
         nombre tal cual, sin inventar ni adivinar.

    Devuelve (nombre_limpio, sede). `sede` puede ser cadena vacía si no se
    detectó ninguna.
    """
    normalized = _normalize_keep_length(raw_name)
    working_name = raw_name
    sede_found = ""

    # --- 1) Coincidencia contra la lista de sedes conocidas ---
    match = _find_known_sede(normalized)
    if match:
        canonical, start, end = match
        sede_found = canonical
        prefix = working_name[:start]
        suffix = working_name[end:]
        # Si la palabra "sede"/"cede" queda pegada justo antes o después,
        # se elimina también para que no quede residual en el nombre.
        prefix = re.sub(r"(sede|cede)\s*[:\-_/#.]*\s*$", "", prefix, flags=re.IGNORECASE)
        suffix = re.sub(r"^\s*[:\-_/#.]*\s*(sede|cede)?\b", "", suffix, flags=re.IGNORECASE)
        working_name = prefix + suffix

    # --- 2) Palabra clave "sede"/"cede" genérica (fallback) ---
    if not sede_found:
        kw_match = _SEDE_KEYWORD_PATTERN.search(raw_name)
        if kw_match:
            after = raw_name[kw_match.end():]
            before = raw_name[:kw_match.start()]
            # Solo se toma como sede el texto que viene DESPUÉS de la
            # palabra clave (ej. "Juan Pérez sede Rengo" -> sede="Rengo").
            # Si no hay letras después (ej. "... sede 14"), no se adivina
            # nada a partir del texto anterior: se deja en blanco.
            after_match = re.match(
                r"\s*[:\-_/#.]*\s*([A-Za-zÀ-ÿ.]+(?:\s+[A-Za-zÀ-ÿ.]+){0,2})", after
            )
            if after_match and after_match.group(1).strip(" ."):
                candidate = after_match.group(1).strip(" .")
                sede_found = candidate.title()
                working_name = (before + after[after_match.end():]).strip()
            else:
                # Se encontró "sede"/"cede" pero sin texto claro asociado;
                # se remueve la palabra para no ensuciar el nombre.
                working_name = (before + after).strip()

    # --- Limpieza final del nombre resultante ---
    # Restos sueltos de la palabra "sede"/"cede" (ej. "sede 043") que hayan
    # quedado pegados al nombre tras una extracción por lista conocida.
    working_name = re.sub(r"\b(sede|cede)\b\s*\d*\s*", " ", working_name, flags=re.IGNORECASE)
    working_name = re.sub(r"\(\s*\)", "", working_name)          # paréntesis vacíos
    working_name = re.sub(r"\[\s*\]", "", working_name)          # corchetes vacíos
    working_name = re.sub(r"\s*[-_/#|]\s*$", "", working_name)   # conectores sueltos al final
    working_name = re.sub(r"^\s*[-_/#|]\s*", "", working_name)   # conectores sueltos al inicio
    working_name = re.sub(r"\s{2,}", " ", working_name).strip()

    return working_name, sede_found


def parse_duration_to_minutes(raw_value: object) -> float:
    """
    Convierte un valor de duración (texto o número) a minutos (float).
    Soporta:
        - Números simples: "45", "45.5", "45,5"
        - Formato hh:mm:ss o mm:ss -> "01:20:00"
        - Formato "1h 20m", "20m", "1h", "1 h 20 min"
    Si no logra interpretarlo, devuelve 0.0 en lugar de fallar.
    """
    if raw_value is None:
        return 0.0
    value = str(raw_value).strip()
    if value == "":
        return 0.0

    # Número simple (admite coma decimal)
    try:
        return float(value.replace(",", "."))
    except ValueError:
        pass

    # Formato hh:mm:ss o mm:ss
    if re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", value):
        parts = [int(p) for p in value.split(":")]
        if len(parts) == 3:
            h, m, s = parts
            return h * 60 + m + s / 60
        m, s = parts
        return m + s / 60

    # Formato "1h 20m", "20m", "1h"
    hours = re.search(r"(\d+)\s*h", value, re.IGNORECASE)
    minutes = re.search(r"(\d+)\s*m", value, re.IGNORECASE)
    if hours or minutes:
        total = 0.0
        if hours:
            total += int(hours.group(1)) * 60
        if minutes:
            total += int(minutes.group(1))
        return total

    return 0.0


def compute_attendance_flag(duration_minutes: float, threshold_minutes: float) -> str:
    """
    Devuelve "P" (Presente) si la duración de asistencia es mayor o igual
    al umbral configurado, o "A" (Ausente) en caso contrario.
    """
    try:
        return "P" if float(duration_minutes) >= float(threshold_minutes) else "A"
    except (TypeError, ValueError):
        return "A"


def load_zoom_csv(filepath: str) -> Tuple[List[Dict[str, object]], Dict[str, Optional[str]]]:
    """
    Carga y normaliza el CSV de asistencia de Zoom.

    Retorna:
        (records, column_map)
        records: lista de diccionarios con llaves
                 "Nombre", "Apellido", "Sede", "Duración (minutos)"
        column_map: columnas originales detectadas (útil para mostrarle
                     al usuario qué columnas se usaron).

    Lanza CSVProcessingError con un mensaje claro si:
        - El archivo no se puede leer / decodificar.
        - No se detectan las columnas mínimas obligatorias.
        - El archivo no contiene filas de datos válidas.
    """
    if _is_disguised_excel_file(filepath):
        raw_text = _load_disguised_excel_as_csv_text(filepath)
    else:
        raw_text = _read_text_with_fallback_encoding(filepath)

    if not raw_text.strip():
        raise CSVProcessingError("El archivo CSV está vacío.")

    delimiter = _sniff_delimiter(raw_text[:2000])
    reader = csv.DictReader(io.StringIO(raw_text), delimiter=delimiter)

    if reader.fieldnames is None:
        raise CSVProcessingError("No se pudieron detectar los encabezados del CSV.")

    headers = [h for h in reader.fieldnames if h is not None and h.strip() != ""]
    column_map = detect_columns(headers)

    missing = []
    if not column_map["full_name"] and not (column_map["first_name"] or column_map["last_name"]):
        missing.append("Nombre del participante")
    if not column_map["duration"]:
        missing.append("Duración de asistencia")

    if missing:
        raise CSVProcessingError(
            "No se encontraron en el archivo las siguientes columnas obligatorias:\n"
            + "\n".join(f"  • {m}" for m in missing)
            + "\n\nEncabezados detectados en el archivo:\n"
            + ", ".join(headers)
            + "\n\nSi Zoom cambió el nombre de estas columnas, se pueden agregar "
              "los nuevos nombres en el archivo csv_processor.py (listas KEYWORDS_*) "
              "sin necesidad de modificar el resto del programa."
        )

    records: List[Dict[str, object]] = []
    for row in reader:
        nombre = ""
        apellido = ""
        sede = ""

        if column_map["first_name"] or column_map["last_name"]:
            nombre = (row.get(column_map["first_name"]) or "").strip() if column_map["first_name"] else ""
            apellido = (row.get(column_map["last_name"]) or "").strip() if column_map["last_name"] else ""
            if not nombre and not apellido and column_map["full_name"]:
                full_name = (row.get(column_map["full_name"]) or "").strip()
                if column_map["sede"]:
                    nombre = full_name
                else:
                    nombre, sede = extract_name_and_sede(full_name)
        elif column_map["full_name"]:
            full_name = (row.get(column_map["full_name"]) or "").strip()
            if column_map["sede"]:
                nombre = full_name
            else:
                nombre, sede = extract_name_and_sede(full_name)

        if column_map["sede"]:
            sede = (row.get(column_map["sede"]) or "").strip()

        duration_raw = row.get(column_map["duration"], "")
        duration_minutes = parse_duration_to_minutes(duration_raw)

        # Se ignoran filas totalmente vacías (Zoom a veces deja líneas en blanco)
        if not nombre and not apellido:
            continue

        records.append({
            "Nombre": nombre,
            "Apellido": apellido,
            "Sede": sede,
            "Duración (minutos)": round(duration_minutes, 2),
            "Asistencia": compute_attendance_flag(duration_minutes, 30.0),
        })

    if not records:
        raise CSVProcessingError(
            "El archivo se leyó correctamente, pero no se encontraron filas "
            "de participantes válidas."
        )

    return records, column_map


def export_records_to_csv(records: List[Dict[str, object]], filepath: str) -> None:
    """Exporta la lista de registros filtrados a un archivo CSV (UTF-8 con BOM,
    compatible con la apertura directa en Excel en español)."""
    columns = ["Nombre", "Apellido", "Sede", "Duración (minutos)", "Asistencia"]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for record in records:
            writer.writerow({col: record.get(col, "") for col in columns})
